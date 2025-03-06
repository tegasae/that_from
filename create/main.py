import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

YEAR='2025'
MONTH='02'

class CreateDate:
    def __init__(self,day:str='',month:str='',year:str=''):
        if day:
            self.day=day
        else:
            self.day=str(datetime.now().day)

        if month:
            self.month=month
        else:
            self.month=str(datetime.now().month)

        if year:
            self.year=year
        else:
            self.year=str(datetime.now().year)
        self.date=datetime.strptime(f'{self.year}-{self.month}-{self.day}', '%Y-%m-%d')

    def year_month(self):
        return self.date.strftime('%Y-%m')

    def year_month_like(self):
        return self.date.strftime('%Y-%m%')

    def month_year_dot(self):
        return self.date.strftime('%m.%Y')

    def get_date(self):
        return self.date.date()


current_date=CreateDate(year=YEAR,month=MONTH)
NAME_XLSX=f'salary-{current_date.year_month()}.xlsx'
@dataclass(kw_only=True)
class Engineer:
    id: int
    code1s: str
    name: str
    remote_hour: float = 0
    remote_salary: float = 0
    duty_days: int = 0
    duty_salary: float = 0
    out_time_hour: float = 0
    out_time_salary: float = 0
    transport: float = 0
    transport_money: float = 0
    fuel:float=0
    fuel_litres: float = 0

@dataclass
class Employee:
    name:str
    hours:float
    hours_fact: float
    code1s: str



@dataclass
class Employee1s(Employee):
    hourly_rate=1400
    @property
    def total(self)->float:
        return self.hours*self.hourly_rate

@dataclass
class Employee1sComplicated(Employee):
    salary=19000
    hourly_rate=2700
    percent=30
    bonus=5000
    @property
    def percent_rub(self)->float:
        return self.hourly_rate*self.hours*(self.percent/100)
    @property
    def total(self) -> float:
        return self.percent_rub+self.salary+self.bonus

@dataclass
class EmployeeWeb(Employee):
    hourly_rate:float=650
    @property
    def total(self) -> float:
        return self.hours_fact*self.hourly_rate


@dataclass
class EmployeeWebComplicated(Employee):
    hourly_rate:float=100
    salary:float=75000
    @property
    def total(self) -> float:
        return self.salary+self.hours*self.hourly_rate





@dataclass
class DateInterval:
    date_start: date
    date_end: date
    count_of_days: int = field(init=False)

    def __init__(self, date_start_str: str, date_end_str: str):
        self.date_start = datetime.strptime(date_start_str, '%Y-%m-%d').date()
        self.date_end = datetime.strptime(date_end_str, '%Y-%m-%d').date()
        self.count_of_days = (self.date_end - self.date_start).days + 1

    def in_month(self, d: date):
        if d.month == self.date_start.month:
            return self.count_of_days
        return 0

    def _date_str(self, date: date) -> str:
        return date.strftime('%Y-%m-%d')

    def start_str(self):
        return self._date_str(self.date_start)

    def end_str(self):
        return self._date_str(self.date_end)

@dataclass
class PersonOil:
    name:str
    address:str
    date:str
    service:str
    operation:str
    quantity:float
    price:float

def get_engineers(conn, engineers: list):
    cur = conn.cursor()
    cur.execute(("SELECT employee_id,employee, code1s FROM employees WHERE department_id=1 and parent='Сотрудники' ORDER BY employee"))
    result = cur.fetchall()
    for i in result:
        engineers.append(Engineer(id=i[0], name=i[1], code1s=i[2]))


def get_remote(conn, enginners: list):
    cur = conn.cursor()
    i = -1
    for e in enginners:
        i += 1
        cur.execute('''select p.employee_id ,sum(p.hours_payment) from performers p 
                        left JOIN  works w on p.work_id =w.work_id 
                        WHERE w.date_ like :date and w.counterparty_id =29 and  p.employee_id =:id
                        group by p.employee_id''', {'date': current_date.year_month_like(), 'id': e.id})
        result = cur.fetchone()
        if not result:
            continue
        enginners[i].remote_hour = result[1]
        s = (result[1] * 70)
        m = s % 1
        s = int(s / 10) * 10
        if m:
            s += 10
        enginners[i].remote_salary = s
def get_out_time(conn, enginners: list):
    cur = conn.cursor()
    i = -1
    for e in enginners:
        i += 1
        cur.execute('''
        select p.employee_id ,sum(p.hours_payment),sum(s.summ)  from performers p 
        left JOIN  works w on p.work_id =w.work_id
        left JOIN services s on w.work_id =s.work_id 
        WHERE w.date_ like :date  and  p.employee_id =:id and s.nomenclature_id=1
        group by p.employee_id  
        ''', {'date': current_date.year_month_like(), 'id': e.id})
        result = cur.fetchone()
        if not result:
            continue

        enginners[i].out_time_hour = result[1]
        enginners[i].out_time_salary = round(result[2] * 0.7)

    i = -1
    for e in enginners:
        i += 1
        cur.execute('''
            select p.employee_id ,sum(p.hours_payment),sum(s.summ)  from performers p 
            left JOIN  works w on p.work_id =w.work_id
            left JOIN services s on w.work_id =s.work_id 
            WHERE w.date_ like :date  and  p.employee_id =:id and s.nomenclature_id=11
            group by p.employee_id  
            ''', {'date': current_date.year_month_like(), 'id': e.id})
        result = cur.fetchone()
        if not result:
            continue

        enginners[i].transport = result[2]
        enginners[i].transport_money = (result[2]/15)*12


def get_duty(conn, enginners: list):
    d_now = current_date.get_date()
    cur = conn.cursor()
    i = -1
    for e in enginners:
        i += 1
        cur.execute('''SELECT date_start,date_end FROM duty_dates WHERE employee_id=:id''', {'id': e.id})
        result = cur.fetchall()
        summ = 0
        if not result:
            continue
        for d in result:
            date_i = DateInterval(date_start_str=d[0], date_end_str=d[1])
            summ += date_i.in_month(d=d_now)

        enginners[i].duty_days=summ
        enginners[i].duty_salary = (summ/7)*1000

def get_fuel(conn, enginners: list):
    d_now = datetime(year=2025, month=9, day=1).date()
    cur = conn.cursor()
    i = -1
    for e in enginners:
        i += 1
        cur.execute('''SELECT sum(ol.quantity),sum(ol.price) FROM oil_logs ol
                        where date_ like :date 
                        and ol.person_id =(select oil_person_id FROM oil_employee oe where oe.employee_id=:id)
                    ''', {'date': current_date.year_month_like(),'id': e.id})
        result = cur.fetchone()

        if not result[0]:
            enginners[i].fuel_litres = 0
            enginners[i].fuel = 0
            continue
        enginners[i].fuel_litres=result[0]
        enginners[i].fuel=result[1]


def get_fuel_all(conn, people: list):
    #d_now = datetime(year=2024, month=9, day=1).date()
    d_now=current_date.get_date()
    cur = conn.cursor()
    i = -1

    cur.execute('''SELECT op.name,ol.date_ ,operation,service,address,quantity,price  from oil_logs ol  
    left join oil_people op on op.person_id =ol.person_id 
    where ol.date_ like :date order by op.name , ol.date_ 
                    ''',{'date': current_date.year_month_like()})
    result = cur.fetchall()
    for r in result:
        people.append(PersonOil(name=r[0],date=r[1],operation=r[2],service=r[3],address=r[4],
                                quantity=r[5],price=r[6]))


def get_hours(conn,employees:list,department_id:int=0):
    cur = conn.cursor()

    cur.execute('''
          select e.employee,sum(p.hours_payment),sum(p.hours_fact),e.code1s  from employees e 
            left join performers p on p.employee_id=e.employee_id
            left join works w 
            on w.work_id =p.work_id 
            where 
            e.department_id =:department_id and e.parent ='Сотрудники' and w.date_ like :date 
            group by e.employee_id,e.code1s 
        ''', {'date': current_date.year_month_like(), 'department_id': department_id})
    result = cur.fetchall()
    i=-1
    for r in result:
        i += 1
        if not result:
            continue
        employees.append(Employee(name=r[0],hours=r[1],hours_fact=r[2],code1s=r[3]))


def get_hours1s(conn):
    employees:list[Employee]=[]
    employees1s=[]
    get_hours(conn=conn,employees=employees,department_id=3)
    for e in employees:
        if e.code1s=='000000087':
            employees1s.append(Employee1sComplicated(name=e.name,hours_fact=e.hours_fact,hours=e.hours,code1s=e.code1s))
        else:
            employees1s.append(Employee1s(name=e.name, hours_fact=e.hours_fact, hours=e.hours, code1s=e.code1s))
    return employees1s


def get_hours_web(conn):
    employees:list[Employee]=[]
    employees_web=[]
    get_hours(conn=conn,employees=employees,department_id=2)
    for e in employees:
        if e.code1s=='000000080':
            employees_web.append(EmployeeWebComplicated(name=e.name,hours_fact=e.hours_fact,hours=e.hours,code1s=e.code1s))
        else:
            employees_web.append(EmployeeWeb(name=e.name, hours_fact=e.hours_fact, hours=e.hours, code1s=e.code1s))
    return employees_web





engineers = []
p1s=[]
webs=[]
people_oil=[]

conn = sqlite3.connect('../1c_work/works.db')

get_engineers(conn, engineers)
get_remote(conn, engineers)
get_out_time(conn, engineers)
get_duty(conn, engineers)
get_fuel(conn,engineers)
#get_hours(conn,p1s,3)
p1s=get_hours1s(conn)
webs=get_hours_web(conn)
get_fuel_all(conn,people_oil)
for e in engineers:
    print(e)

print(p1s)
print(webs)

conn.close()



wb = Workbook()

wb.save(NAME_XLSX)
ws = wb.active
ws.title='Расчет часов и расходов'
ws.row_dimensions[5].height = 50
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 15

ws['A3'] = current_date.month_year_dot()
ws['A3'].font=Font(bold=True)
ws['A4'] = 'Суппорт'
ws['A4'].font = Font(bold=True)
row=5

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


ws.cell(row=5, column=1).value='Имя'
ws.cell(row=5, column=1).font=Font(bold=True)
ws.cell(row=5, column=1).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=1).border=thin_border

ws.cell(row=5, column=2).value='Удаленка, часы'
ws.cell(row=5, column=2).font=Font(bold=True)
ws.cell(row=5, column=2).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=2).border=thin_border

ws.cell(row=5, column=3).value='Удаленка, рубли'
ws.cell(row=5, column=3).font=Font(bold=True)
ws.cell(row=5, column=3).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=3).border=thin_border


ws.cell(row=5, column=4).value='Дежурство, дни'
ws.cell(row=5, column=4).font=Font(bold=True)
ws.cell(row=5, column=4).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=4).border=thin_border

ws.cell(row=5, column=5).value='Дежурство, рубли'
ws.cell(row=5, column=5).font=Font(bold=True)
ws.cell(row=5, column=5).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=5).border=thin_border

ws.cell(row=5, column=6).value='Внеурочное, часы'
ws.cell(row=5, column=6).font=Font(bold=True)
ws.cell(row=5, column=6).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=6).border=thin_border

ws.cell(row=5, column=7).value='Внеурочное, рубли'
ws.cell(row=5, column=7).font=Font(bold=True)
ws.cell(row=5, column=7).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=7).border=thin_border

ws.cell(row=5, column=8).value='Транспортные, рубли'
ws.cell(row=5, column=8).font=Font(bold=True)
ws.cell(row=5, column=8).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=8).border=thin_border

ws.cell(row=5, column=9).value='Транспортные на руки, рубли'
ws.cell(row=5, column=9).font=Font(bold=True)
ws.cell(row=5, column=9).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=9).border=thin_border

ws.cell(row=5, column=10).value='Бензин, литры'
ws.cell(row=5, column=10).font=Font(bold=True)
ws.cell(row=5, column=10).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=10).border=thin_border

ws.cell(row=5, column=11).value='Бензин, рубли'
ws.cell(row=5, column=11).font=Font(bold=True)
ws.cell(row=5, column=11).alignment=Alignment(wrapText=True)
ws.cell(row=5, column=11).border=thin_border

for e in engineers:
    row+=1
    ws.cell(row=row,column=1).value=e.name
    ws.cell(row=row, column=1).border = thin_border

    ws.cell(row=row, column=2).value = e.remote_hour
    ws.cell(row=row, column=2).border = thin_border

    ws.cell(row=row, column=3).value = e.remote_salary
    ws.cell(row=row, column=3).border = thin_border

    ws.cell(row=row, column=4).value=e.duty_days
    ws.cell(row=row, column=4).border = thin_border

    ws.cell(row=row, column=5).value = e.duty_salary
    ws.cell(row=row, column=5).border = thin_border

    ws.cell(row=row, column=6).value = e.out_time_hour
    ws.cell(row=row, column=6).border = thin_border

    ws.cell(row=row, column=7).value = e.out_time_salary
    ws.cell(row=row, column=7).border = thin_border

    ws.cell(row=row, column=8).value = e.transport
    ws.cell(row=row, column=8).border = thin_border

    ws.cell(row=row, column=9).value = e.transport_money
    ws.cell(row=row, column=9).border = thin_border

    ws.cell(row=row, column=10).value = e.fuel_litres
    ws.cell(row=row, column=10).border = thin_border

    ws.cell(row=row, column=11).value = e.fuel
    ws.cell(row=row, column=11).border = thin_border


ws.cell(row=7+len(engineers),column=1).value='1C'
ws.cell(row=7+len(engineers),column=1).font=Font(bold=True)
ws.cell(row=7+len(engineers)+1,column=1).value='Имя'
ws.cell(row=7+len(engineers)+1,column=1).font=Font(bold=True)
ws.cell(row=7+len(engineers)+1,column=2).value='Часы на оплату'
ws.cell(row=7+len(engineers)+1,column=2).font=Font(bold=True)
ws.cell(row=7+len(engineers)+1,column=3).value='Часы, факт'
ws.cell(row=7+len(engineers)+1,column=3).font=Font(bold=True)
ws.cell(row=7+len(engineers)+1,column=4).value='Оплата'
ws.cell(row=7+len(engineers)+1,column=4).font=Font(bold=True)



row=7+len(engineers)+1
for e in p1s:
    row+=1
    ws.cell(row=row, column=1).value = e.name
    ws.cell(row=row, column=2).value = e.hours_fact
    ws.cell(row=row, column=3).value = e.hours
    ws.cell(row=row, column=4).value = e.total
    if type(e) is Employee1sComplicated:
        ws.cell(row=row,column=5).value='Оклад:'
        ws.cell(row=row, column=6).value = e.salary
        ws.cell(row=row, column=7).value = 'Ставка:'
        ws.cell(row=row, column=8).value = e.hourly_rate
        ws.cell(row=row, column=9).value = 'Премия:'
        ws.cell(row=row, column=10).value = e.bonus
        ws.cell(row=row, column=11).value = 'Процент:'
        ws.cell(row=row, column=12).value = e.percent

row=7+len(engineers)+len(p1s)+4
ws.cell(row=row,column=1).value='Веб'
ws.cell(row=row,column=1).font=Font(bold=True)
row+=1
ws.cell(row=row,column=1).value='Имя'
ws.cell(row=row,column=1).font=Font(bold=True)
ws.cell(row=row,column=2).value='Часы на оплату'
ws.cell(row=row,column=2).font=Font(bold=True)
ws.cell(row=row,column=3).value='Часы, факт'
ws.cell(row=row,column=3).font=Font(bold=True)
ws.cell(row=row,column=4).value='Сумма'
ws.cell(row=row,column=4).font=Font(bold=True)



for e in webs:
    row+=1
    ws.cell(row=row, column=1).value = e.name
    ws.cell(row=row, column=2).value = e.hours
    ws.cell(row=row, column=3).value = e.hours_fact
    ws.cell(row=row, column=4).value = e.total

ws1=wb.create_sheet('Бензин')


ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 30
#ws.column_dimensions['D'].width = 20
#ws.column_dimensions['E'].width = 20
#ws.column_dimensions['F'].width = 20
#ws.column_dimensions['G'].width = 20
#ws.column_dimensions['H'].width = 20
#ws.column_dimensions['I'].width = 20
#ws.column_dimensions['J'].width = 20


name=''
position=3
ws1.cell(row=position, column=2).value = 'Дата'
ws1.cell(row=position, column=2).font = Font(bold=True)
ws1.cell(row=position, column=3).value = 'Адрес'
ws1.cell(row=position, column=3).font = Font(bold=True)
ws1.cell(row=position, column=4).value = 'Услуга'
ws1.cell(row=position, column=4).font = Font(bold=True)
ws1.cell(row=position, column=5).value = 'Операция'
ws1.cell(row=position, column=5).font = Font(bold=True)
ws1.cell(row=position, column=6).value = 'Количество'
ws1.cell(row=position, column=6).font = Font(bold=True)
ws1.cell(row=position, column=7).value = 'Цена'
ws1.cell(row=position, column=7).font = Font(bold=True)


position+=1
sum_quantity=0
sum_price=0
sum_quantity_full=0
sum_price_full=0


for i in range(1,len(people_oil)+1):
    if name!=people_oil[i-1].name:
        if name!='':

            ws1.cell(row=position, column=6).value = sum_quantity
            ws1.cell(row=position, column=6).font=Font(bold=True)
            ws1.cell(row=position, column=7).value = sum_price
            ws1.cell(row=position, column=7).font = Font(bold=True)
            position+=1

        position+=1
        ws1.cell(row=position,column=1).value=people_oil[i-1].name
        name=people_oil[i-1].name



        sum_price = 0
        sum_quantity=0


    ws1.cell(row=position, column=2).value = people_oil[i - 1].date
    ws1.cell(row=position, column=3).value = people_oil[i - 1].address
    ws1.cell(row=position, column=4).value = people_oil[i - 1].service
    ws1.cell(row=position, column=5).value = people_oil[i - 1].operation
    ws1.cell(row=position, column=6).value = people_oil[i - 1].quantity
    ws1.cell(row=position, column=7).value = people_oil[i - 1].price

    sum_quantity+=people_oil[i - 1].quantity
    sum_price+=people_oil[i - 1].price

    sum_quantity_full = sum_quantity_full + people_oil[i - 1].quantity
    sum_price_full = sum_price_full + people_oil[i - 1].price

    position+=1

ws1.cell(row=position, column=6).value = sum_quantity
ws1.cell(row=position, column=6).font=Font(bold=True)
ws1.cell(row=position, column=7).value = sum_price
ws1.cell(row=position, column=7).font = Font(bold=True)
position+=1

position+=1
ws1.cell(row=position, column=6).value =  sum_quantity_full
ws1.cell(row=position, column=6).font=Font(bold=True)
ws1.cell(row=position, column=7).value = sum_price_full
ws1.cell(row=position, column=7).font = Font(bold=True)



wb.close()
wb.save(NAME_XLSX)

