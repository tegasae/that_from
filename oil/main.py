import sqlite3

import openpyxl

import pyexcel as p
import re

from models.domain import Log, Card, PersonOil
from repository.card import SqlLiteRepositoryPerson

xls_file = 'oil_january.xls'
xlsx_file = 'oil_january.xlsx'

# Convert XLS to XLSX
p.save_book_as(file_name=xls_file, dest_file_name=xlsx_file)

dataframe = openpyxl.load_workbook("oil_january.xlsx")
datasheet = dataframe.worksheets[0]

people: dict[str, PersonOil] = {}

empty_lines = 0
i = 1
while empty_lines < 30:
    card_number = datasheet.cell(row=i, column=2).value
    name = datasheet.cell(row=i, column=5).value
    i += 1

    if type(card_number) is not str or card_number=='':
        empty_lines += 1
        continue
    if re.match('^\d{5,}', card_number):
        empty_lines=0
        if type(name) is str:
            p = PersonOil(name=name)
            if not name in people.keys():
                people[p.name] = p

            if not card_number in people[p.name].cards.keys():
                people[p.name].cards[card_number] = Card(number=card_number)

        while True:

            date = datasheet.cell(row=i, column=3).value
            if not date:
                break
            # date = datetime.strptime(date, '%d.%m.%Y %H:%M:%S')
            address = datasheet.cell(row=i, column=5).value
            operation = datasheet.cell(row=i, column=6).value
            service = datasheet.cell(row=i, column=7).value
            quanity = datasheet.cell(row=i, column=8).value
            price = datasheet.cell(row=i, column=9).value
            log = Log(date_log=date, address=address, operation=operation, service=service, quantity=quanity,
                      price=price)

            # print(p)
            if not log in people[p.name].cards[card_number].logs:
                people[p.name].cards[card_number].logs.append(log)

            i += 1
        people[p.name].cards[card_number].check_summ=datasheet.cell(row=i+3, column=9).value
#        if not people[p.name].cards[card_number].checking_summ():
#            raise ValueError(f"Data isn't valid {p.name} {card_number}")

con = sqlite3.connect("../1c_work/works.db")
rp=SqlLiteRepositoryPerson(conn=con)
for i in people.keys():
    print(rp.add(people[i]))
    # print(f"{people[i]} {people[i].cards}")
for i in rp.people1c.keys():
    print(rp.people1c[i].short_name())
    if rp.people1c[i].short_name() in rp.seen.keys():
        print("True")
con.commit()
con.close()


