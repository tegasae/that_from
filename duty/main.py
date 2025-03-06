import json
import sqlite3
from dataclasses import dataclass, field, asdict
from datetime import date, datetime

import re

import openpyxl


def get_cell_background_color(cell):
    fill = cell.fill

    # Check if the cell has a fill pattern
    if fill.patternType is None:
        return None  # No fill applied

    fg_color = fill.fgColor

    # If the color is in RGB format
    if fg_color.type == 'rgb':
        return fg_color.rgb  # Returns a string like 'FFFFFFFF' (ARGB format)

    # If the color is indexed (uses a standard color index)
    elif fg_color.type == 'indexed':
        # Map the indexed color to RGB (requires a lookup)
        index = fg_color.indexed
        # OpenPyXL provides a standard color index mapping
        from openpyxl.styles.colors import COLOR_INDEX
        rgb = COLOR_INDEX[index]
        return rgb

    # If the color is based on a theme
    elif fg_color.type == 'theme':
        # Theme colors are more complex; they depend on the workbook's theme settings
        theme = fg_color.theme
        tint = fg_color.tint  # This is a float value indicating the tint
        # To convert theme and tint to RGB, you need to access the workbook's theme elements
        # This process is complex and may require additional code or libraries
        return f"Theme color: {theme}, Tint: {tint}"

    else:
        return None  # Unknown color type


start_row = 3
start_col = 3
current_year = "2025"


@dataclass
class DateInterval:
    date_start: date
    date_end: date
    count_of_days: int = field(init=False)

    def __init__(self, date_str: str, current_year: str):
        digits = re.match(r"(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})\D+(\d{1,2})", date_str)

        self.date_start = date(year=int(current_year), day=int(digits[1]), month=int(digits[2]))
        self.date_end = date(year=int(current_year), day=int(digits[3]), month=int(digits[4]))
        self.count_of_days = (self.date_end - self.date_start).days + 1

    def in_month(self, date: datetime):
        if date.month == self.date_start.month:
            return self.count_of_days
        return 0

    def _date_str(self, date: date) -> str:
        return date.strftime('%Y-%m-%d')

    def start_str(self):
        return self._date_str(self.date_start)

    def end_str(self):
        return self._date_str(self.date_end)


@dataclass
class Engineer:
    name: str
    code: str
    dates: list[DateInterval] = field(default_factory=list)

    def get_days(self, date: datetime):
        summ = 0
        for i in self.dates:
            summ += i.in_month(date)
        return summ


def default_serializer(obj):
    if isinstance(obj, date):
        return obj.isoformat()  # Convert date to ISO format string
    else:
        return str(obj)
    # raise TypeError(f"Type {type(obj)} not serializable")


engineers: list[Engineer] = []
date_intervals: list[DateInterval] = []

dataframe = openpyxl.load_workbook("Дежурства 2025.xlsx")
datasheet = dataframe.worksheets[0]
# for row in range(0, datasheet.max_row):
#    for col in datasheet.iter_cols(1, datasheet.max_column):
#        fill = col[row].fill
#        if fill.patternType is None:
#            f = "Not fill"
#        else:
#            f = "fill"
#        print(f"{col[row].coordinate} {col[row].value} {get_cell_background_color(col[row])}")
#
#        # print(col[row].value)
#        # print(col[row].fill.bgColor.value)
#    print("------")
print(datasheet.cell(start_row, start_col).value)
##Получаем инхенеров
for i in range(start_row + 1, datasheet.max_row + 1):
    name = datasheet.cell(i, start_col).value
    code = datasheet.cell(i, start_col - 1).value
    print(i)
    print(name)
    if name is None:
        break
        # raise ValueError("The value mustn't be None")
    engineers.append(Engineer(name=name, code=code))

##Получаем даты
for i in range(start_col + 1, datasheet.max_column + 1):
    v = datasheet.cell(start_row, i).value
    if v is None:
        #raise ValueError("The value mustn't be None")
        print("The value mustn't be None")
        continue
    date_intervals.append(DateInterval(v, current_year))

for r in range(start_row + 1, datasheet.max_row + 1):
    for c in range(start_col + 1, datasheet.max_column + 1):
        # print(engineers[r-start_row-1])
        # print(f"[{r},{c}]={datasheet.cell(r,c).value} ",end="")
        if not datasheet.cell(3,c).value:
            print("Date is null")
            continue
        if get_cell_background_color(datasheet.cell(r, c)):
            # print(engineers[r - start_row - 1])
            # print(date_intervals[c - start_col - 1])
            # print(datasheet.cell(r, c))
            engineers[r - start_row - 1].dates.append(date_intervals[c - start_col - 1])

# print(engineers)
# print(date_intervals)

j = []
for i in engineers:
    j.append(asdict(i))

# print(json.dumps(j,default=default_serializer,ensure_ascii=False))
for i in engineers:
    print(i)
date_format = "%Y-%m-%d"
date_start = datetime.strptime('2025-01-01', date_format)
date_end = datetime.strptime('2025-01-31', date_format)

conn = sqlite3.connect('../1c_work/works.db')
cur = conn.cursor()
for i in engineers:
    cur.execute("SELECT employee_id FROM employees WHERE code1s=:code1s", {'code1s': i.code})
    result = cur.fetchone()
    cur.execute("DELETE FROM duty_dates WHERE employee_id=:employee_id", {'employee_id': result[0]})

    for d in i.dates:

        cur.execute("INSERT INTO duty_dates (date_start,date_end,employee_id) "
                    "VALUES(:date_start,:date_end,:employee_id)",
                    {'date_start':d.start_str(), 'date_end': d.end_str(), 'employee_id': result[0]})


    print(i.name)
    print(i.get_days(date_start))
conn.commit()
conn.close()
